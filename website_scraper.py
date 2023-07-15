import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook

url = "https://fbref.com/en/comps/9/Premier-League-Stats"
response = requests.get(url)
html_content = response.text

# Parse the HTML
soup = BeautifulSoup(html_content, "html.parser")

# Find the table
table = soup.find("table")

if table:
    # Extract the class or ID of the table
    table_class = table.get("class")
    table_id = table.get("id")

    print("Table class:", table_class)
    print("Table ID:", table_id)


    # Extract player data
    player_data = []

    # Iterate over each row in the table
    rows = table.find_all("tr")[:1]
    for row in rows:
        # Extract player name and rating
        name = row.find("td", class_="Squad").text.strip()
        rating = row.find("td", class_="GF").text.strip()

        # Append to list
        player_data.append((name, rating))

    # Create the excel file using Workbook
    wb = Workbook()
    sheet = wb.active

    # Write the corresponding name to the file
    for i, (name, rating) in enumerate(player_data, start=1):
        sheet.cell(row=i, column=1, value=name)
        sheet.cell(row=i, column=2, value=rating)

    wb.save("First_Test.xlsx")
    print("Excel file saved successfully.")
else:
    print("Table not found.")




